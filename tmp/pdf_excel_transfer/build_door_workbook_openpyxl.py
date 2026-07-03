import re
from copy import copy
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


FLOOR_MARKER = "\u0e23\u0e30\u0e14\u0e31\u0e1a\u0e0a\u0e31\u0e49\u0e19"
ROOF_PREFIX = "\u0e14\u0e32\u0e14\u0e1f"
ROOF_LABEL = "\u0e14\u0e32\u0e14\u0e1f\u0e49\u0e32"

DOOR_ID_RE = re.compile(r"^\s*([A-Z0-9]+(?:-[A-Z0-9]+)+)\b")
SIZE_RE = re.compile(r"\b\d{3,4}x\d{3,4}\b")
DOOR_CODE_RE = re.compile(r"^(?:AD|D)[A-Za-z0-9.]*$")
CODE_SORT_RE = re.compile(r"^(AD|D)(\d+)([A-Z]?)(?:\.(\d+))?(?:f(\d+))?$")
CODE_FALLBACK_RE = re.compile(r"^(AD|D)(\d+)(.*)$")


def natural_key(value: str) -> Tuple[Tuple[int, object], ...]:
    parts = re.findall(r"\d+|\D+", str(value))
    key = []
    for part in parts:
        if part.isdigit():
            key.append((0, int(part)))
        else:
            key.append((1, part))
    return tuple(key)


def door_code_sort_key(code: str) -> Tuple[object, ...]:
    if not code:
        return (9, 999, "", 9, 999, 9, 999, "")

    detailed = CODE_SORT_RE.match(code)
    if detailed:
        prefix, number, letter, dot_part, fire_part = detailed.groups()
        return (
            0 if prefix == "D" else 1,
            int(number),
            letter or "",
            0 if dot_part is None else 1,
            int(dot_part or 0),
            0 if fire_part is None else 1,
            int(fire_part or 0),
            code,
        )

    fallback = CODE_FALLBACK_RE.match(code)
    if fallback:
        prefix, number, suffix = fallback.groups()
        return (0 if prefix == "D" else 1, int(number), suffix, 9, 999, 9, 999, code)

    return (8, 999, code, 9, 999, 9, 999, code)


def parse_pdf_rows(pdf_text: str) -> Tuple[List[str], List[Dict[str, object]]]:
    floor_order: List[str] = []
    records: List[Dict[str, object]] = []
    current_floor = None
    original_index = 0

    for raw_line in pdf_text.splitlines():
        if FLOOR_MARKER in raw_line:
            tail = raw_line.split(FLOOR_MARKER, 1)[1].strip()
            if tail.startswith(ROOF_PREFIX):
                current_floor = ROOF_LABEL
            elif tail:
                current_floor = tail.split()[0]
            if current_floor and current_floor not in floor_order:
                floor_order.append(current_floor)
            continue

        door_match = DOOR_ID_RE.match(raw_line)
        if not current_floor or not door_match:
            continue

        raw_window = raw_line[:800]
        size_match = SIZE_RE.search(raw_window)
        if not size_match:
            continue

        after_size = raw_window[size_match.end() :]
        margin_match = re.search(r"\s{80,}\S", after_size)
        table_line = (
            raw_window[: size_match.end() + margin_match.start()]
            if margin_match
            else raw_window
        )

        left_side = table_line[: size_match.start()].rstrip()
        tokens = left_side.strip().split()
        maybe_code = tokens[-1] if len(tokens) >= 2 else ""
        door_code = maybe_code if DOOR_CODE_RE.match(maybe_code) else ""

        equipment_candidates = re.findall(r"\b\d+\b", table_line[size_match.end() :])
        equipment = equipment_candidates[-1] if equipment_candidates else ""

        records.append(
            {
                "floor": current_floor,
                "door_id": door_match.group(1),
                "door_code": door_code,
                "equipment": int(equipment) if equipment else None,
                "original_index": original_index,
            }
        )
        original_index += 1

    return floor_order, records


def build_output_rows(
    floor_order: List[str], records: List[Dict[str, object]]
) -> List[Dict[str, object]]:
    by_floor: Dict[str, List[Dict[str, object]]] = {floor: [] for floor in floor_order}
    for record in records:
        by_floor.setdefault(record["floor"], []).append(record)

    output_rows: List[Dict[str, object]] = []
    for floor in floor_order:
        items = sorted(
            by_floor.get(floor, []),
            key=lambda item: (
                door_code_sort_key(str(item["door_code"] or "")),
                natural_key(str(item["door_id"])),
                item["original_index"],
            ),
        )

        previous_code = None
        first_row = True
        for item in items:
            if not first_row and item["door_code"] != previous_code:
                output_rows.append({"kind": "blank"})

            output_rows.append(
                {
                    "kind": "data",
                    "values": [
                        floor if first_row else None,
                        item["door_code"] or None,
                        item["door_id"],
                        item["equipment"],
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                }
            )

            previous_code = item["door_code"]
            first_row = False

        output_rows.append({"kind": "separator"})

    return output_rows


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 9) -> None:
    for col in range(1, max_col + 1):
        src_cell = ws.cell(src_row, col)
        dst_cell = ws.cell(dst_row, col)
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)
        if src_cell.number_format:
            dst_cell.number_format = copy(src_cell.number_format)
        if src_cell.font:
            dst_cell.font = copy(src_cell.font)
        if src_cell.fill:
            dst_cell.fill = copy(src_cell.fill)
        if src_cell.border:
            dst_cell.border = copy(src_cell.border)
        if src_cell.alignment:
            dst_cell.alignment = copy(src_cell.alignment)
        if src_cell.protection:
            dst_cell.protection = copy(src_cell.protection)

    src_dim = ws.row_dimensions[src_row]
    dst_dim = ws.row_dimensions[dst_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden
    dst_dim.outlineLevel = src_dim.outlineLevel


def write_row_values(ws, row_number: int, values: List[object]) -> None:
    for col, value in enumerate(values, start=1):
        ws.cell(row_number, col).value = value


def clear_row_values(ws, row_number: int, max_col: int = 9) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row_number, col).value = None


def main() -> None:
    import sys

    if len(sys.argv) != 4:
        raise SystemExit(
            "Usage: build_door_workbook_openpyxl.py <input.xlsx> <pdf_text.txt> <output.xlsx>"
        )

    input_xlsx = Path(sys.argv[1])
    pdf_text_path = Path(sys.argv[2])
    output_xlsx = Path(sys.argv[3])

    floor_order, records = parse_pdf_rows(pdf_text_path.read_text(encoding="utf-8"))
    output_rows = build_output_rows(floor_order, records)

    workbook = load_workbook(input_xlsx)
    ws = workbook.active

    data_template_row = 2
    blank_template_row = 5
    separator_template_row = 25

    row_number = 2
    for row in output_rows:
        if row["kind"] == "data":
            copy_row_style(ws, data_template_row, row_number)
            write_row_values(ws, row_number, row["values"])
        elif row["kind"] == "blank":
            copy_row_style(ws, blank_template_row, row_number)
            clear_row_values(ws, row_number)
        else:
            copy_row_style(ws, separator_template_row, row_number)
            clear_row_values(ws, row_number)
        row_number += 1

    for clear_row in range(row_number, max(ws.max_row, row_number) + 1):
        clear_row_values(ws, clear_row)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)

    print(
        {
            "floors": floor_order,
            "recordCount": len(records),
            "outputRowCount": len(output_rows) + 1,
            "outputPath": str(output_xlsx),
        }
    )


if __name__ == "__main__":
    main()
